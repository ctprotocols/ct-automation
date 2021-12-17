# -*- coding: utf-8 -*-
"""
Created on Thu Aug 13 07:03:00 2020

@author: EastmanE
"""
# =============================================================================
# If the program is stopped before line 100 or runs into an error before the copying
# is complete, run the command 'excel.Quit'. Then, open task manager and quit
# Excel. 
# =============================================================================
#####Import packages
import datetime as dt
import openpyxl as xl
import pandas as pd
import os
import win32com.client as win32
from CTQC_settings_2021 import chartpath, chartnames, box_path, Z_path
import copy_last_month_2021

# =============================================================================
# Set Dates
# =============================================================================
#Yesterday's date yyyy-mm-dd 00:00:00
yesterday = (dt.datetime.today()-dt.timedelta(days=1)).replace(hour = 0, minute = 0, second = 0, microsecond = 0)   #Yesterday's date yyyy-mm-dd 00:00:00

#Month name in 3 letter format (e.g. Aug)   
month_shortname= yesterday.strftime('%b')    
#Year in YYYY format            
year = yesterday.strftime('%Y')                
#YYYY-mm          
year_month = yesterday.strftime('%Y-%m')        
#YYYY-mm August        
year_month_monthname = yesterday.strftime('%Y-%m %B')   

# =============================================================================
# Dispatch Excel and create the destination file
# =============================================================================
#Destination folder name in Z drive
destination_folder = year_month + ' CT QC'                     
#Destination file name. This is what the program will name the file it creates.
destination_file = 'CT QC ' + year_month + '.xlsx'          
#Full path to the destination
destination_path = os.path.join(Z_path, destination_folder, destination_file)    

#Start Excel
excel = win32.Dispatch('Excel.Application')              
#Excel creates a new workbook
CTQC_wb = excel.Workbooks.Add()       
#Turns off any prompts from Excel (e.g. do you want to overwrite existing file?)
excel.DisplayAlerts = False       
#Saves the new workbook to the path. 
CTQC_wb.SaveAs(destination_path)               

# =============================================================================
# Find paths
# =============================================================================
#Empty list to hold paths
box_paths_list = []                                      
#Look for directories in Box QA folder
for QA_dir in os.listdir(box_path):                             
    #Add to the path
    path_L1 = os.path.join(box_path, QA_dir)                    
    #If it is a folder and not a file, List directories in each machine's folder
    if not os.path.isfile(path_L1):                             
        for QA_dir_L2 in os.listdir(path_L1):
            #Create the paths to each directory
            path_L2 = os.path.join(path_L1, QA_dir_L2)          
            #If the path leads to a file and the YYYY_MM is in the name, Add to the list (Most machines have this format)
            if (os.path.isfile(path_L2)) & (year_month in QA_dir_L2):   
                box_paths_list.append(path_L2)                          
            #If the path leads to a folder and YYYY is in the folder name (Machine > 2020 folder > file)
            elif (not os.path.isfile(path_L2)) & (year in QA_dir_L2): 
                #List directories in the YYYY folder
                for QA_dir_L3 in os.listdir(path_L2):
                    #Find the YYYY_MM in the file name (MGB)
                    if year_month in QA_dir_L3:                         
                        #Create path, append path
                        path_L3 = os.path.join(path_L2, QA_dir_L3)      
                        box_paths_list.append(path_L3)                  
            #Find the AHSP file and append it
            elif (os.path.isfile(path_L2)) & (year in QA_dir_L2) & ('AHSP' in QA_dir_L2):   
                box_paths_list.append(path_L2)
            else:
                pass
    else:
        pass

# =============================================================================
# Copy QC worksheets
# =============================================================================
print('Copying QC sheets...')
for path in box_paths_list:
    #open the worksheet
    machine_wb = excel.Workbooks.Open(Filename = path)                
    #AHSP worksheetname = 3 letters of the month all caps
    if 'AHSP' in path:
        machine_ws = machine_wb.Worksheets(month_shortname.upper())
    #All others are name "QA"
    else:
        machine_ws = machine_wb.Worksheets('QA')                                   
    #Copy the worksheet into the new workbook
    machine_ws.Copy(Before= CTQC_wb.Worksheets(1))             
    #Rename the sheet in the new workbook
    new_sheetname = path.split(os.sep)[6]
    CTQC_wb.Worksheets(1).Name = new_sheetname
    #Close the original worksheet and don't save any changes
    machine_wb.Close(SaveChanges=False)        
    #Print name to show progress
    print(new_sheetname, 'QC copied.')        
    
# =============================================================================
# Copy Linearity worksheets
# =============================================================================
#create a list of paths for Tosh/Canon machines (they have Lin tests as well)
lin_paths = [path for i, path in enumerate(box_paths_list) if ('Toshiba' in path) | ('Canon' in path)]

for i, lin_file_path in enumerate(lin_paths):
    #Open up the workbook
    lin_wb = excel.Workbooks.Open(Filename = lin_file_path)                      
    #Navigate to the lineariy worksheet, worksheet in workbook must be named "Toshiba Linearity"
    lin_ws = lin_wb.Worksheets('Toshiba Linearity')                       
    #Copy and rename
    lin_ws.Copy(Before= CTQC_wb.Worksheets(1))                               
    lin_name = lin_file_path.split(os.sep)[6]
    CTQC_wb.Worksheets(1).Name = lin_name + ' Linearity'                 
    #Close the workbook from box and don't save any changes to original workbook
    lin_wb.Close(SaveChanges=False)                                      
    print(lin_name, 'Linearity copied.')
# =============================================================================
# Copy summary graphs from the template into the spreadsheet
# =============================================================================
#Open workbook containing templates for the three charts
chartwb = excel.Workbooks.Open(Filename = chartpath) 

#Copy the three charts. Chart names are defined in the settings file 
for names in chartnames:                                
    chartws = chartwb.Worksheets(names)                 
    chartws.Copy(Before= CTQC_wb.Worksheets(1))    
    #Change the link for these worksheets to the current worksheet.         
    CTQC_wb.ChangeLink(Name = chartpath, NewName= destination_path)    
    #Rename the sheets
    CTQC_wb.Worksheets(1).Name = names                      
    print(names, 'Chart copied.')






# =============================================================================
# Save/Close/Quit
# =============================================================================
#Delete Sheet1. This sheet is automatically created when you first create the file.     
CTQC_wb.Worksheets('Sheet1').Delete()   
CTQC_wb.Close(SaveChanges=True)         
#Quit Excel. If the program gets stopped before this line, run this command to quit Excel.
excel.Quit()                          
print('Copying complete.')

# =============================================================================
# Look for daily fails
# =============================================================================
print('Looking for fails and warnings...')
#Name of file that will hold the whole month's notifications
summary_file = 'Summary Results ' + year_month + '.csv'
#Name of file that will hold today's notifications
results_today_file = 'Todays Results ' + year_month + '.csv'
#Create paths for the two files above
summary_path = os.path.join(Z_path, destination_folder, summary_file)    
results_today_path = os.path.join(Z_path, destination_folder, results_today_file)   

#Use openpyxl to load the workbook we just created
des_wb = xl.load_workbook(destination_path, data_only=True) 

 

#Results will be added to this string to form the body of the email.
Emailbody = ''                                                      
#Dataframes to collect results in 
concatdf = pd.DataFrame()
blankconcatdf = pd.DataFrame()

#Create list of machine names

machine_names = [path.split(os.sep)[6] for path in box_paths_list]


#Go through each machine's sheet
for i, name in enumerate(machine_names):                          
    #Load destination worksheet
    des_ws = des_wb[name] 
    #Convert Excel data to pandas DataFrame                                   
    df=pd.DataFrame(des_ws.values)                     
    #Drop additional columns to avoid counting duplicate warnings/fails
    if 'MGB' in name:
        df.drop(df.columns[16:], axis=1, inplace=True)  
    else:
        df.drop(df.columns[22:], axis=1, inplace=True)  
    #Filter out CLOSED and DOWN machines (while keeping monthly results)    
    open_df=df.loc[(df[1] != 'CLOSED') & (df[1] != 'WARNING') & (df[2] != 'FAIL') & (df[2] != 'WARNING')]   
    
    #Create a dataframe to collect the warning or failures
    warning_df = pd.DataFrame(columns = open_df.columns)
    fail_df = pd.DataFrame(columns = open_df.columns)
    #Find fail/warn for open machines
    for j in open_df.columns:                                  
        warning_df= pd.concat([warning_df, open_df.loc[open_df[j] == 'WARNING']])
        fail_df= pd.concat([fail_df, open_df.loc[open_df[j] == 'FAIL']])
    
    # =============================================================================
    # Fill in any blank dates with "Monthly Check" or the date from the row above (for dual tubes)
    # =============================================================================
    #index values to list
    w_index_vals = warning_df.index.tolist()       
    f_index_vals = fail_df.index.tolist()       
    
    #Check if it is tube B of a dual tube, then fill in the missing date
    for k, w_idx in enumerate(w_index_vals):         
        try:
            if 'B' in warning_df.loc[w_idx,1]:                               
                #Pull the date from the row above (tube A)
                warning_df.loc[w_idx,0] = df.loc[w_index_vals[k]-1, 0]       
            else:
                pass
        #If it encounters an error (Nonetype), pass
        except TypeError:                                           
            pass
    for m, f_idx in enumerate(f_index_vals):
        try:
            if 'B' in fail_df.loc[f_idx,1]:                 
                fail_df.loc[f_idx,0] = df.loc[f_index_vals[m]-1, 0]
            else:
                pass
        except TypeError:
            pass
    
    #Remaining blanks in the date column are for monthly tests. 
    warning_df.loc[warning_df[0].isnull(), 0] = 'Monthly Check'  
    fail_df.loc[fail_df[0].isnull(), 0] = 'Monthly Check'
    
    # =============================================================================
    # Add machine name  and results. Concatenate the dataframes.
    # =============================================================================
    warning_df.rename(columns={0: 'Date'}, inplace= True)
    warning_df.insert(1, 'Result', 'WARNING')
    warning_df.insert(1, 'Machine', name)

    fail_df.rename(columns={0: 'Date'}, inplace= True)
    fail_df.insert(1, 'Result', 'FAIL')
    fail_df.insert(1, 'Machine', name)
    
    failandwarndf = pd.concat([warning_df, fail_df])
    
    failtypelist = []
    for row in failandwarndf.index:
        for col in failandwarndf.columns:
            if (failandwarndf.loc[row,col] == 'FAIL') | (failandwarndf.loc[row,col] == 'WARNING'):
                #We are looking for the original appearance of fail/warning. We expect there to be a fail/warning in the Results column
                if col == 'Result':         
                    continue                
                #If the failure is a monthly one
                elif failandwarndf.loc[row, 'Date'] == 'Monthly Check':   
                    if ('AHSP' in name) | ('Force' in name):
                        failtype = df.loc[row, 4]
                        failtypelist.append(failtype)
                    else:
                        failtype = df.loc[row, 3]
                        failtypelist.append(failtype)
                else:
                    failtype = df.iloc[4, (col-1)]
                    failtypelist.append(failtype)
            else:
                pass
    failandwarndf.insert(2, 'Test', failtypelist)
    concatdf = pd.concat([concatdf, failandwarndf.iloc[:, :4]], ignore_index=True)

# =============================================================================
# Check for Linearity daily failures and warnings
# =============================================================================
lin_sheets = [name.split(os.sep)[6] + ' Linearity' for i, name in enumerate(lin_paths)]

for sheet in lin_sheets:                                        
    lin_ws = des_wb[sheet]                       
    lin_df = pd.DataFrame(lin_ws.values)                           
    lin_df.drop(lin_df.columns[13:], axis=1, inplace = True)    

    #Create empty dataframe to collect failures    
    lin_fail_df = pd.DataFrame()                                          
    #This will only keep the row if it fails/is blank and is open
    linfail = lin_df.loc[(lin_df[9] == 'FAIL') & (lin_df[1] == 'OPEN')]       
    lin_fail_df = lin_fail_df.append(linfail) 

    #Empty dataframe = no fails.    
    if not lin_fail_df.empty:                                 
        lin_fail_df.rename(columns={0: 'Date'}, inplace= True)
        lin_fail_df.insert(1, 'Result', 'FAIL')    
        lin_fail_df.insert(1, 'Test', 'Linearity')
        lin_fail_df.insert(1, 'Machine', sheet)
        #Add linearity fails
        concatdf = pd.concat([concatdf, lin_fail_df.iloc[:, :4]]).reset_index(drop = True)
    else:
        pass

# #Print results
# if not concatdf.empty:
#     print('Daily/monthly warning/fail found.')
# else:
#     print('No daily/monthly warning/fail found.')
        
# =============================================================================
# Reformat the results data
# =============================================================================
for row in concatdf.index:          
    try:
        #Remove the time from the datetime format
        concatdf.loc[row, 'Date'] = dt.datetime.strftime(concatdf.loc[row, 'Date'], '%Y-%m-%d') 
    #If it's a string/if you can't perform the function, leave it alone
    except TypeError:   
        try: 
            concatdf.loc[row, 'Date']= dt.datetime.strftime(dt.datetime.strptime(concatdf.loc[row, 'Date'], '%m/%d/%Y'), '%Y-%m-%d')
        except TypeError:   
            pass
#Sort by date
concatdf = concatdf.sort_values(by = ['Date'], ascending = True).reset_index(drop = True) 

# =============================================================================
# Compare results to results from the summary file
# =============================================================================
#Use try in case it is the first day of the month and there is no file existing yet.
try:                                        
    comparedf = pd.read_csv(summary_path)      
    #Reformat the dates if they were converted. Sometimes the dates are changes from yyyy-mm-dd to mm/dd/yyyy
    for row in comparedf.index:             
        try:
            #Read in datetimes in a different format
            comparedf.loc[row, 'Date'] = dt.datetime.strptime(comparedf.loc[row, 'Date'], '%m/%d/%Y') 
        except ValueError:
            pass
        try:
            #Reformat into desired format
            comparedf.loc[row, 'Date'] = dt.datetime.strftime(comparedf.loc[row, 'Date'], '%Y-%m-%d') 
        except TypeError:
            pass
    #Concatenate today's results and yesterday's. Remove duplicates.
    newconcatdf = pd.concat([concatdf, comparedf]).drop_duplicates().reset_index(drop = True)
    #Sort by date
    newconcatdf = newconcatdf.sort_values(by = ['Date'], ascending = True).reset_index(drop = True)
    #Write the results to csv file
    newconcatdf.to_csv(summary_path, index = False)                                 

    # =============================================================================
    # Pull only new alerts
    # =============================================================================
    comparison_df = comparedf.merge(newconcatdf, indicator = True, how = 'outer')
    new_alerts_df = comparison_df.loc[comparison_df['_merge'] == 'right_only']
    new_alerts_df = new_alerts_df.iloc[:, :4]
    new_alerts_df.to_csv(results_today_path, index = False)

# =============================================================================
# If the file doesn't exist, it is the first day of the month. 
# Nothing to compare it to so just write today's data to the file.
# =============================================================================
except FileNotFoundError:           
    #Rename the dataframe
    newconcatdf = concatdf                                                                         
    #Sort by date
    newconcatdf = newconcatdf.sort_values(by = ['Date'], ascending = True).reset_index(drop = True) 
    #Write the results to csv files
    newconcatdf.to_csv(summary_path, index = False)                                                    
    newconcatdf.to_csv(results_today_path, index = False)
    new_alerts_df = newconcatdf                                                                                          


#If there is a failure found today, add it to the email body
if not new_alerts_df.empty:       
    Emailbody += 'New Fails/Warnings\n' + new_alerts_df.to_html( index = False)  
    print('Fail or warning found.')

else:
    print('Congratulations! No fail or warning found.')                 




# =============================================================================
# Look for blank SMPTE, Artifact, and Daily Checks from the prev week on Thursdays
# =============================================================================
#Blank check occurs on Thursday. This is date of the Sunday of the week prior
blank_start = (yesterday - dt.timedelta(days = 10)).replace(hour = 0, minute = 0, second = 0, microsecond = 0)     
#Blank check. This is the date of the Saturday of the week prior. 
blank_stop = (yesterday - dt.timedelta(days = 4)).replace(hour = 0, minute = 0, second = 0, microsecond = 0)        

#Returns the day of the week where Mon = 0 and Sunday = 6. Blank check will run on Thursday (weekday = 3)
weekday = dt.datetime.today().weekday()                 

if weekday == 3:
    print('Checking for blanks...')
    if blank_start.month != yesterday.month:
        copy_last_month_2021.copy_function(blank_start)
        print('Last month QC copied.')
    else:
        pass

    for i, name in enumerate(machine_names):
        blank_df = pd.DataFrame()
        
        #If the start of the time period is last month
        if blank_start.month != yesterday.month:
            #The start of the previous month
            month_start = blank_start.replace(day = 1, hour = 0, minute = 0, second = 0, microsecond = 0)
            #If the start and stop are both in last month
            if blank_stop.month != yesterday.month:                                          
                #Pull data from the destination folder from last month
                destination_folder = blank_start.strftime('%Y-%m') + ' CT QC'                             
                destination_file = 'CT QC ' + blank_start.strftime('%Y-%m') + '.xlsx'          
                lastmonpath = os.path.join(Z_path, destination_folder, destination_file)    
                #Load the workbook
                lastmonth_wb = xl.load_workbook(lastmonpath, read_only = True, data_only=True)                    
                lastmonth_ws = lastmonth_wb[name]                                    
                #Load all of the data into a DataFrame
                month_df = pd.DataFrame(lastmonth_ws.values)                                 
                lastmonth_wb.close()
                month_df.drop(index = 4, inplace = True)
                #csv for last month's blanks
                blank_csv_file = 'Blank Results ' + blank_start.strftime('%Y-%m') + '.csv'         
                # print(name)
            #If the start is last month but the stop is this month
            else:                                                                   
                #Load this month's data
                des_wb = xl.load_workbook(destination_path, read_only = True, data_only = True)
                thismonth_ws = des_wb[name]                                    
                month_df = pd.DataFrame(thismonth_ws.values)      
                des_wb.close()                            
                month_df.drop(index = 4, inplace = True)
                #Load last month's data
                destination_folder = blank_start.strftime('%Y-%m') + ' CT QC'                             
                destination_file = 'CT QC ' + blank_start.strftime('%Y-%m') + '.xlsx'          
                lastmonpath = os.path.join(Z_path, destination_folder, destination_file)    
                lastmonth_wb = xl.load_workbook(lastmonpath, read_only = True, data_only=True)                    
                lastmonth_ws = lastmonth_wb[name]                                    
                last_week_df = pd.DataFrame(lastmonth_ws.values)  
                lastmonth_wb.close()
                last_week_df = last_week_df.drop(index = 4)
                #Combine last month's data to this month's data
                month_df = pd.concat([last_week_df, month_df]).reset_index(drop = True)
                #csv for last month's blanks
                blank_csv_file = 'Blank Results ' + blank_start.strftime('%Y-%m') + '.csv'       
                #csv for this month's blanks
                blank_csv_file_thismonth = 'Blank Results ' + year_month + '.csv'          
                # print(name)
        #If the start and stop are both in this month
        else: 
            #the start of this month
            month_start = yesterday.replace(day = 1, hour = 0, minute = 0, second = 0, microsecond = 0)
            #Load this month's data
            des_wb = xl.load_workbook(destination_path, read_only = True, data_only = True)
            thismonth_ws = des_wb[name]
            month_df = pd.DataFrame(thismonth_ws.values)
            des_wb.close()                            
            month_df.drop(index = 4, inplace = True)
            #csv for this month's blanks
            blank_csv_file = 'Blank Results ' + year_month + '.csv'          
            # print(name)
        
        #Reformat the dates if they were converted. Sometimes the dates are changed from yyyy-mm-dd to mm/dd/yyyy    
        for row in month_df.index:             
            try:
                #Read in datetimes in a different format
                month_df.loc[row, 0] = dt.datetime.strptime(month_df.loc[row, 0], '%m/%d/%Y') 
            except TypeError:                   
                pass
            except ValueError:                  
                pass
        
        # =============================================================================
        # Go through month_df and isolate just the rows with dates we need        
        # =============================================================================
        #Find row containing the start of last month
        dfstart = month_df.loc[month_df[0] == month_start]    
        startidx = dfstart.index[0]                     

        #Row containing the start date for the weekly check (Sunday)
        smptedfstart = month_df.loc[month_df[0] == blank_start]       
        smpte_startidx = smptedfstart.index[0]

        #last day to check this month
        dfstop = month_df.loc[month_df[0] == blank_stop]              
        #Index of the stop row
        stopidx = dfstop.index[0] +1                       

        #If the start date is in the last month and the stop date date is in this month 
        if (blank_start.month != yesterday.month) & (blank_stop.month == yesterday.month):
            #Last day of last month
            midstop = blank_stop.replace(day = 1)-dt.timedelta(days = 1)    
            #Row containing end of last month
            dfmidstop = month_df.loc[month_df[0] == midstop]
            #adjust the end of the month index for dual tube machines that have an additional row below the date
            if ('AHSP' in name) | ('Force' in name):
                midstopidx = dfmidstop.index[0] + 2
            else:
                midstopidx = dfmidstop.index[0] +1

            #first day of this month
            midstart = blank_stop.replace(day = 1)
            #Row containing start of this month
            dfmidstart = month_df.loc[month_df[0] == midstart]
            #if there are two occurences of the first day of this month
            #The excel table in the QC worksheet has 31 date cells. If there are only 30
            #days in the month, the last cell will have the first day of the current month
            #and cause an error
            #Row containing the end date (Saturday)
            #We don't want to look at the entire current month because
            #we give techs a one week window after to fill in 
            if dfmidstart.shape[0] >1:
                #remove the first entry (the one from the previous month)
                dfmidstart = dfmidstart[1:]
            else:
                pass
            midstartidx = dfmidstart.index[0]

            
            #Isolate the data from the whole period, fills in any blank dates 
            #and appends each row in between the start and stop date
            for idx in list(range(startidx, midstopidx)) + list(range(midstartidx, stopidx)):            
                idxrow = month_df.loc[idx].copy()                   
                #If it is a dual tube, column 0 will be missing a date 
                if idxrow[0] == None:                             
                    #If date is missing, replace with date from the row above
                    idxrow[0] = month_df.loc[idx-1, 0]       
                #If date is not missing, skip this section
                else:
                    pass                                    
                #Append the row. 
                blank_df = blank_df.append(idxrow)          
            
        
        #If the whole week falls in either the previous month
        #or the current month
        else:     
            for idx in range(startidx, stopidx):            #This section fills in any blank dates and appends each row in between the start and stop date
                idxrow = month_df.loc[idx].copy()                   #Look at each date's row one at a time
                if idxrow[0] == None:                       #If it is a dual tube, column 0 will be missing a date       
                    idxrow[0] = month_df.loc[idx-1, 0]       #If date is missing, replace with date from the row above
                else:
                    pass                                    #If date is not missing, skip this section
                blank_df = blank_df.append(idxrow)          #Append the row. 
        
        smpte_blank_df = blank_df
        blank_df = blank_df.loc[(blank_df[1] != 'CLOSED') & (blank_df[1] != 'DOWN') & (blank_df[2] != 'CLOSED') & (blank_df[2] != 'DOWN')]  #Filter out any dates where the machine was closed. 
        # =============================================================================
        # This section tells the code which columns to look in and fills them into the dataframes dailydf, artifactdf, and SMPTEdf
        # =============================================================================
        
        if ('Angeles' in name) | ('BOLD' in name):
            statusdf = blank_df.loc[:, 0:1]
            dailydf = blank_df.loc[:, 6:14]
            artifactdf = blank_df.loc[:, 3]
            SMPTEdf=smpte_blank_df.loc[:, 15:17]
        elif ('Revolution' in name):
            statusdf = blank_df.loc[:, 0:1]
            dailydf = blank_df.loc[:, 6:14]
            dailydf = dailydf.drop([9,10], axis = 1)
            artifactdf = blank_df.loc[:, 3]
            SMPTEdf=smpte_blank_df.loc[:, 15:17]
        elif ('AHSP' in name) | ('Force' in name):
            smpte_dualdf = smpte_blank_df.loc[:, 4:18]
            smpte_dualdf_idx = smpte_dualdf.index.tolist()
            
            dualdf = blank_df.loc[:, 4:18]
            dualdf_idx = dualdf.index.tolist()
            
            for row in range(0, len(dualdf_idx)-1):
                currentrow = dualdf_idx[row]
                nextrow = dualdf_idx[row +1]
                if (dualdf.loc[currentrow, 7] == 'Helical') & (dualdf.loc[nextrow, 7] == 'Axial'): #If it is a helical scan, only tube A is used. Remove the second helical row because it will give a false alert since it empty.
                    dualdf = dualdf.drop(currentrow)
                else:
                    pass
            for row in range(0, len(smpte_dualdf_idx)-1):
                currentrow = smpte_dualdf_idx[row]
                nextrow = smpte_dualdf_idx[row +1]
                if (smpte_dualdf.loc[currentrow, 7] == 'Helical') & (smpte_dualdf.loc[nextrow, 7] == 'Axial'): #If it is a helical scan, only tube A is used. Remove the second helical row because it will give a false alert since it empty.
                    smpte_dualdf = smpte_dualdf.drop(currentrow)
                else:
                    pass

            statusdf = dualdf.loc[:, 0:2]
            dailydf = dualdf.loc[:, 7:15]
            artifactdf = dualdf.loc[:, 4]             
            SMPTEdf=smpte_dualdf.loc[:, 16:18]                       
        elif 'MGB' in name:
            statusdf = blank_df.loc[:, 0:1]
            dailydf = blank_df.loc[:, 6:11]
            artifactdf = blank_df.loc[:, 3]
            SMPTEdf=smpte_blank_df.loc[:, 12].to_frame()
        else:
            statusdf = blank_df.loc[:, 0:1]
            dailydf = blank_df.loc[:, 6:12]
            artifactdf = blank_df.loc[:, 3]
            SMPTEdf=smpte_blank_df.loc[:, 13:15]                       
        
        
        # =============================================================================
        # This section tells the code which columns to look in and fills them into the dataframes dailydf, artifactdf, and SMPTEdf
        # =============================================================================
        if dailydf.isnull().any().any():    #If any of the cells are blank in this section, create alert. All cells should have a value.
            dailyconcat = pd.DataFrame()    
            dailyidx = dailydf[dailydf.isnull().any(axis = 1)].index.tolist()
            for idx in dailyidx:
                dailyconcat = dailyconcat.append(blank_df.loc[idx, :])     #Append warning rows for open machines to empty df
            dailyconcat.rename(columns={0: 'Date'}, inplace= True)
            dailyconcat.insert(1, 'Result', 'Blank')     #Append warning rows for open machines to empty df
            dailyconcat.insert(1, 'Test', 'Daily CT # Check')
            dailyconcat.insert(1, 'Machine', name)
            blankconcatdf = pd.concat([blankconcatdf, dailyconcat]).reset_index(drop = True)
        else:
            pass
        
        if statusdf.isnull().any().any():
            statusconcat = pd.DataFrame()
    
            if ('AHSP' in name) | ('Force' in name):
                statusidx = statusdf[statusdf[2].isnull()].index.tolist()
                for idx in statusidx:
                    statusconcat = statusconcat.append(blank_df.loc[idx, :])     #Append warning rows for open machines to empty df
            else:
                statusidx = statusdf[statusdf[1].isnull()].index.tolist()
                for idx in statusidx:
                    statusconcat = statusconcat.append(blank_df.loc[idx, :])     #Append warning rows for open machines to empty df
            statusconcat.rename(columns={0: 'Date'}, inplace= True)
            statusconcat.insert(1, 'Result', 'Blank')     #Append warning rows for open machines to empty df
            statusconcat.insert(1, 'Test', 'Unit Status')
            statusconcat.insert(1, 'Machine', name)
            blankconcatdf = pd.concat([blankconcatdf, statusconcat]).reset_index(drop = True)
        else:
            pass
            
        if artifactdf.isnull().any().any(): #If any of the cells are blank in this section, create alert. All cells should have a value.
            artifactconcat = pd.DataFrame() 
            artifactidx = artifactdf[artifactdf.isnull()].index.tolist()
            for idx in artifactidx:
                artifactconcat = artifactconcat.append(blank_df.loc[idx, :])     #Append warning rows for open machines to empty df
            artifactconcat.rename(columns={0: 'Date'}, inplace= True)
            artifactconcat.insert(1, 'Result', 'Blank')     #Append warning rows for open machines to empty df
            artifactconcat.insert(1, 'Test', 'Free of Artifacts')
            artifactconcat.insert(1, 'Machine', name)
            blankconcatdf = pd.concat([blankconcatdf, artifactconcat]).reset_index(drop = True)
        else:
            pass
        
        # =============================================================================
        # Check for weekly test
        # =============================================================================
        if ('AHSP' in name) | ('Force' in name):
            SMPTEdays = [smpte_blank_df.loc[row, 0].day for row in smpte_dualdf.index]
            SMPTEdates = [smpte_blank_df.loc[row,0] for row in smpte_dualdf.index]
            SMPTEdf.insert(0, 'date', SMPTEdates)
            SMPTEdf.insert(0, 'weekday', SMPTEdays)
        else:
            SMPTEdays = [i.day for i in smpte_blank_df.loc[:, 0]]
            SMPTEdf.insert(0, 'date', smpte_blank_df.loc[:, 0])
            SMPTEdf.insert(0, 'weekday', SMPTEdays)
        # for row in SMPTEdf.index:
        #     if SMPTEdf.loc[row, 'weekday'] != 0:
        #         SMPTEdf.drop(row, inplace = True)
        #     else:
        #         break
        if ('AHSP' in name) | ('Force' in name):
            droplist = []
            for dupidx, duprow in enumerate(SMPTEdf.index[:-1]):
                if SMPTEdf.loc[duprow, 'weekday'] == SMPTEdf.loc[SMPTEdf.index[dupidx+1], 'weekday']:
                    if (SMPTEdf.iloc[dupidx, 2:].isnull().all()) & (SMPTEdf.iloc[dupidx+1, 2:].isnull().all()):
                        droplist.append(duprow)
                    elif (not SMPTEdf.iloc[dupidx, 2:].isnull().all()) & (SMPTEdf.iloc[dupidx+1, 2:].isnull().all()):
                        droplist.append(SMPTEdf.index[dupidx+1])
                    elif (SMPTEdf.iloc[dupidx, 2:].isnull().all()) & (not SMPTEdf.iloc[dupidx+1, 2:].isnull().all()):
                        droplist.append(duprow)
                    else:
                        droplist.append(duprow)
            SMPTEdf.drop(droplist, axis = 0, inplace = True)
        
        
        smpteconcat = pd.DataFrame() 
        dayofweek = 0
        for idx, row in enumerate(SMPTEdf.index):
            if dayofweek == 0:
                count = 0
            if not SMPTEdf.iloc[idx, 2:].isnull().all():
                count += 1
            dayofweek +=1
            if dayofweek == 7:
                if count == 0:
                    smpterow = pd.DataFrame() 
                    smptedate = SMPTEdf.loc[SMPTEdf.index[idx-6], 'date']
                    smpterow = smpterow.append({'Date': smptedate, 'Machine': name, 'Test' : 'Weekly SMPTE', 'Result' : 'Blank'}, ignore_index=True)
                    smpteconcat = pd.concat([smpteconcat, smpterow])
                    dayofweek = 0
                else:
                    dayofweek =0 
        smpteconcat.drop_duplicates(inplace = True)
                    
        blankconcatdf = pd.concat([blankconcatdf, smpteconcat]).reset_index(drop = True)
     
    # =============================================================================
    # Reformat the blankconcatdf                  
    # =============================================================================
    blankconcatdf = blankconcatdf[['Date', 'Machine', 'Test', 'Result']]     #trim excess columns
    
    for row in blankconcatdf.index:          #Reformat the dates
        try:       
            blankconcatdf.loc[row, 'Date'] = dt.datetime.strftime(blankconcatdf.loc[row, 'Date'], '%Y-%m-%d' )
        except TypeError:
            pass
    
    if (blank_start.month != yesterday.month) & (blank_stop.month == yesterday.month):
        blankconcatdf = blankconcatdf.sort_values(by = ['Date'], ascending = [True]).reset_index(drop = True) #Sort by date
        
        if dt.datetime.strptime(blankconcatdf.iloc[-1,0], '%Y-%m-%d') <= midstop:
            prevblankconcatdf = blankconcatdf
            thisblankconcatdf = pd.DataFrame()
        elif dt.datetime.strptime(blankconcatdf.iloc[0,0], '%Y-%m-%d') >= midstart:
            prevblankconcatdf = pd.DataFrame()
            thisblankconcatdf = blankconcatdf

        else:
            for row in blankconcatdf.index:
                try:
                    if (dt.datetime.strptime(blankconcatdf.iloc[row,0], '%Y-%m-%d') <= midstop) & (dt.datetime.strptime(blankconcatdf.iloc[row+1,0], '%Y-%m-%d') >= midstart):
                        prevblankconcatdf = blankconcatdf.iloc[:row+1, :]
                        thisblankconcatdf = blankconcatdf.iloc[row+1:, :]
                    else:
                        pass
                except ValueError:
                    pass
                except IndexError:
                    pass
        csv_path2 = os.path.join(Z_path, destination_folder, blank_csv_file)   #Current blanks path
        try:
            prevblankconcatdf = prevblankconcatdf.sort_values(by = ['Machine', 'Date'], ascending = [True, True]).reset_index(drop = True) #Sort by date
            prevblankconcatdf.to_csv(csv_path2, index = False)
            Emailbody += '\nCurrent Blanks (last month)\n ' + prevblankconcatdf.to_html( index = False)
        except NameError:
            prevblankconcatdf = pd.DataFrame(columns = ['Date', 'Machine', 'Test', 'Result'])
            prevblankconcatdf.to_csv(csv_path2, index = False)

        
    ###    csvname4 = 'Blank Results ' + year_month + '.csv'          
        csv_path4 = os.path.join(Z_path, year_month + ' CT QC', blank_csv_file_thismonth)   #Current blanks path
        if not thisblankconcatdf.empty:
            try:
                thisblankconcatdf = thisblankconcatdf.sort_values(by = ['Machine', 'Date'], ascending = [True, True]).reset_index(drop = True) #Sort by date
                thisblankconcatdf.to_csv(csv_path4, index = False)
                Emailbody += '\nCurrent Blanks\n' + thisblankconcatdf.to_html( index = False)
            except NameError:
                thisblankconcatdf = blankconcatdf.sort_values(by = ['Machine', 'Date'], ascending = [True, True]).reset_index(drop = True) #Sort by date
                thisblankconcatdf.to_csv(csv_path4, index = False)
                Emailbody += '\nCurrent Blanks\n' + thisblankconcatdf.to_html( index = False)

        
        # if prevblankconcatdf.empty != True:       
        #     lastmonth_wb = xl.load_workbook(lastmonpath, data_only=True)                    #Uses openpyxl to load the workbook. data_only to read values and not formulas
        #     writer = pd.ExcelWriter(lastmonpath, engine = 'openpyxl')                 #Prep to write summary dataframe into existing Excel doc
        #     writer.book = lastmonth_wb
        #     writer.sheets = dict((ws.title, ws) for ws in lastmonth_wb.worksheets)
        #     prevblankconcatdf.to_excel(writer, 'Current Blanks', index = False)
        #     writer.save()
        #     writer.close()

        # if thisblankconcatdf.empty != True:       
        #     lastmonth_wb = xl.load_workbook(destination_path, data_only=True)                    #Uses openpyxl to load the workbook. data_only to read values and not formulas
        #     writer = pd.ExcelWriter(destination_path, engine = 'openpyxl')                 #Prep to write summary dataframe into existing Excel doc
        #     writer.book = lastmonth_wb
        #     writer.sheets = dict((ws.title, ws) for ws in lastmonth_wb.worksheets)
        #     thisblankconcatdf.to_excel(writer, 'Current Blanks', index = False)
        #     writer.save()
        #     writer.close()
    else: 
        csv_path2 = os.path.join(Z_path, destination_folder, blank_csv_file)   #Current blanks path
        blankconcatdf = blankconcatdf.sort_values(by = ['Machine', 'Date'], ascending = [True, True]).reset_index(drop = True) #Sort by date
        blankconcatdf.to_csv(csv_path2, index = False)
        Emailbody += '\nCurrent Blanks\n' + blankconcatdf.to_html( index = False)

        lastmonth_wb = xl.load_workbook(destination_path, data_only=True)                    #Uses openpyxl to load the workbook. data_only to read values and not formulas
        # writer = pd.ExcelWriter(destination_path, engine = 'openpyxl')                 #Prep to write summary dataframe into existing Excel doc
        # writer.book = lastmonth_wb
        # writer.sheets = dict((ws.title, ws) for ws in lastmonth_wb.worksheets)
        # blankconcatdf.to_excel(writer, 'Current Blanks', index = False)

        # writer.save()
        # writer.close()

    if not blankconcatdf.empty:
        print('Blanks found.')
    else:
        print('Congratulations! No blanks found.')
# =============================================================================
# Send Email
# =============================================================================
if not new_alerts_df.empty:         #If the email is blank (no failures or warnings are found), don't send an email.
    subject_date = (dt.datetime.today()).strftime('%Y-%m-%d')  #Set date for the subject
    outlook = win32.Dispatch('outlook.application')     #Launch outlook
    mail = outlook.CreateItem(0)                        #Create a mail item
    mail.To = 'emi.eastman@cshs.org; Yifang.Zhou@cshs.org; Alexander.Scott@cshs.org'                  #Define recipients
    # mail.To = 'emi.eastman@cshs.org'                  #Define recipients
    mail.Subject = 'Daily CT QC Alert: ' + subject_date #Define subject
    mail.HTMLBody = Emailbody                               #Define body
    attachment1  = results_today_path                              #Add an attachment
    mail.Attachments.Add(attachment1)
    if (weekday == 3) & (not blankconcatdf.empty):
        if (blank_start.month != yesterday.month) & (blank_stop.month == yesterday.month):
            attachment2  = csv_path2                              #Add an attachment
            mail.Attachments.Add(attachment2)
            if not thisblankconcatdf.empty:
                attachment3  = csv_path4                              #Add an attachment
                mail.Attachments.Add(attachment3)
        else:
            attachment2  = csv_path2                              #Add an attachment
            mail.Attachments.Add(attachment2)
        mail.Send()                                         #Send email
        print('Email alert sent.')
    else:
        mail.Send()                                         #Send email
        print('Email alert sent.')
else:
    if (weekday == 3) & (not blankconcatdf.empty):
        subject_date = (dt.datetime.today()).strftime('%Y-%m-%d')  #Set date for the subject
        outlook = win32.Dispatch('outlook.application')     #Laun ch outlook
        mail = outlook.CreateItem(0)                        #Create a mail item
        mail.To = 'emi.eastman@cshs.org; Yifang.Zhou@cshs.org; Alexander.Scott@cshs.org'                  #Define recipients
        # mail.To = 'emi.eastman@cshs.org'                  #Define recipients
        mail.Subject = 'Daily CT QC Alert: ' + subject_date #Define subject
        mail.HTMLBody = Emailbody                               #Define body
        if (blank_start.month != yesterday.month) & (blank_stop.month == yesterday.month):
            attachment2  = csv_path2                              #Add an attachment
            mail.Attachments.Add(attachment2)
            if not thisblankconcatdf.empty:
                attachment3  = csv_path4                              #Add an attachment
                mail.Attachments.Add(attachment3)

        else:
            attachment2  = csv_path2                              #Add an attachment
            mail.Attachments.Add(attachment2)
        mail.Send() 
        print('Email alert sent.')                                        #Send email
    else:
        pass
print('Program is complete.')
